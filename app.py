import io
import sqlite3
import pandas as pd
from fastapi import FastAPI
from fastapi.responses import HTMLResponse, Response

from cache import (
    cache_data,
    get_cached_data
)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = FastAPI(title="Smart Web Scraper System")


def get_conn():
    return sqlite3.connect("data/scraped_data.db")

def fetch_df(query, params=None):
    conn = get_conn()
    df = pd.read_sql(query, conn, params=params)
    conn.close()
    return df

SOURCE_PILLS = {
    "Hacker News":    "pill-hn",
    "Reddit":         "pill-rd",
    "The Hindu":      "pill-th",
    "Economic Times": "pill-et",
    "YourStory":      "pill-ys",
}

SOURCE_COLORS_XLSX = {
    "Hacker News":    "FF6400",
    "Reddit":         "FF4500",
    "The Hindu":      "00ADB5",
    "Economic Times": "7EE787",
    "YourStory":      "F78166",
}


@app.get("/", response_class=HTMLResponse)
async def home():

    df = fetch_df("""
        SELECT title, link, source, scraped_at
        FROM articles
        ORDER BY scraped_at DESC
        LIMIT 100
    """)

    rows_html = ""
    for _, row in df.iterrows():
        pill_cls = SOURCE_PILLS.get(row['source'], 'pill-def')
        rows_html += f"""
            <tr>
                <td><a href="{row['link']}" target="_blank" title="{row['title']}">{row['title']}</a></td>
                <td><span class="source-pill {pill_cls}">{row['source']}</span></td>
                <td><span class="ts">{row['scraped_at']}</span></td>
            </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Smart Web Scraper Dashboard</title>
        <link href="https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
        <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
        <style>

            :root {{
                --bg:       #080c10;
                --surface:  #0d1117;
                --surface2: #161b22;
                --border:   #21262d;
                --accent:   #00d9e8;
                --accent2:  #7ee787;
                --warn:     #f78166;
                --text:     #e6edf3;
                --muted:    #7d8590;
                --mono:     'Space Mono', monospace;
                --body:     'DM Sans', sans-serif;
            }}

            *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

            body {{
                background: var(--bg);
                color: var(--text);
                font-family: var(--body);
                font-size: 14px;
                min-height: 100vh;
                overflow-x: hidden;
            }}

            body::before {{
                content: '';
                position: fixed; inset: 0;
                background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.04'/%3E%3C/svg%3E");
                pointer-events: none; z-index: 0;
            }}

            body::after {{
                content: '';
                position: fixed; inset: 0;
                background-image:
                    linear-gradient(rgba(0,217,232,.03) 1px, transparent 1px),
                    linear-gradient(90deg, rgba(0,217,232,.03) 1px, transparent 1px);
                background-size: 40px 40px;
                pointer-events: none; z-index: 0;
            }}

            .shell {{
                position: relative; z-index: 1;
                display: grid;
                grid-template-columns: 240px 1fr;
                min-height: 100vh;
            }}

            /* ── Sidebar ── */
            .sidebar {{
                background: var(--surface);
                border-right: 1px solid var(--border);
                padding: 28px 20px;
                position: sticky; top: 0;
                height: 100vh;
                display: flex; flex-direction: column; gap: 4px;
                overflow-y: auto;
            }}

            .sidebar-logo {{
                font-family: var(--mono);
                font-size: 13px; color: var(--accent);
                letter-spacing: .08em;
                padding: 0 8px 24px;
                border-bottom: 1px solid var(--border);
                margin-bottom: 16px;
                display: flex; align-items: center; gap: 10px;
            }}

            .logo-dot {{
                width: 8px; height: 8px;
                background: var(--accent2); border-radius: 50%;
                animation: pulse 2s ease-in-out infinite; flex-shrink: 0;
            }}

            @keyframes pulse {{
                0%,100% {{ opacity:1; transform:scale(1); }}
                50%      {{ opacity:.4; transform:scale(.7); }}
            }}

            .nav-label {{
                font-family: var(--mono); font-size: 10px;
                color: var(--muted); letter-spacing: .12em;
                text-transform: uppercase;
                padding: 12px 8px 4px;
            }}

            .nav-link {{
                display: flex; align-items: center; gap: 10px;
                padding: 9px 12px; border-radius: 8px;
                color: var(--muted); text-decoration: none;
                font-size: 13px; font-weight: 500;
                transition: all .15s;
                border: 1px solid transparent;
            }}

            .nav-link:hover, .nav-link.active {{
                color: var(--text); background: var(--surface2);
                border-color: var(--border);
            }}

            .nav-link .icon {{ font-size: 15px; width: 20px; text-align: center; }}

            .sidebar-footer {{
                margin-top: auto; padding-top: 16px;
                border-top: 1px solid var(--border);
            }}

            .refresh-badge {{
                display: flex; align-items: center; gap: 8px;
                font-family: var(--mono); font-size: 10px;
                color: var(--muted); padding: 8px 12px;
            }}

            .refresh-dot {{
                width: 6px; height: 6px;
                background: var(--accent2); border-radius: 50%;
                animation: pulse 3s ease-in-out infinite;
            }}

            /* ── Main ── */
            .main {{
                padding: 32px 36px;
                display: flex; flex-direction: column; gap: 28px;
                overflow: hidden;
            }}

            @keyframes fadeDown {{
                from {{ opacity:0; transform:translateY(-12px); }}
                to   {{ opacity:1; transform:translateY(0); }}
            }}

            /* ── Header ── */
            .header {{
                display: flex; align-items: flex-start;
                justify-content: space-between;
                gap: 20px; flex-wrap: wrap;
                animation: fadeDown .4s ease both;
            }}

            .page-title {{
                font-family: var(--mono); font-size: 26px;
                font-weight: 700; line-height: 1.1;
            }}

            .page-title span {{ color: var(--accent); }}
            .page-sub {{ font-size: 13px; color: var(--muted); margin-top: 4px; font-weight: 300; }}

            #clock {{
                font-family: var(--mono); font-size: 12px; color: var(--muted);
                background: var(--surface2); border: 1px solid var(--border);
                padding: 8px 14px; border-radius: 8px; white-space: nowrap;
            }}

            /* ── Stat cards ── */
            .stats-grid {{
                display: grid;
                grid-template-columns: repeat(3, 1fr);
                gap: 16px;
                animation: fadeDown .4s .1s ease both;
            }}

            .stat-card {{
                background: var(--surface); border: 1px solid var(--border);
                border-radius: 12px; padding: 20px 22px;
                position: relative; overflow: hidden;
                transition: border-color .2s, transform .2s;
            }}

            .stat-card::before {{
                content: '';
                position: absolute; top: 0; left: 0; right: 0; height: 2px;
                background: var(--accent); opacity: 0; transition: opacity .2s;
            }}

            .stat-card:hover {{ border-color: var(--accent); transform: translateY(-3px); }}
            .stat-card:hover::before {{ opacity: 1; }}

            .stat-label {{
                font-family: var(--mono); font-size: 10px;
                letter-spacing: .1em; text-transform: uppercase;
                color: var(--muted); margin-bottom: 10px;
            }}

            .stat-value {{
                font-family: var(--mono); font-size: 32px;
                font-weight: 700; line-height: 1;
            }}

            .stat-value.accent {{ color: var(--accent); }}
            .stat-value.green  {{ color: var(--accent2); }}
            .stat-value.small  {{ font-size: 18px; margin-top: 4px; }}
            .stat-icon {{ position: absolute; top: 18px; right: 18px; font-size: 22px; opacity: .2; }}

            /* ── Analytics card ── */
            .analytics-card {{
                background: var(--surface); border: 1px solid var(--border);
                border-radius: 12px; padding: 24px;
                animation: fadeDown .4s .15s ease both;
            }}

            .analytics-title {{
                font-family: var(--mono); font-size: 11px;
                letter-spacing: .1em; text-transform: uppercase;
                color: var(--muted); margin-bottom: 20px;
            }}

            /* Two charts side-by-side, SAME height */
            .charts-row {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 20px;
            }}

            /* White rounded box that constrains each chart */
            .chart-box {{
                background: #ffffff;
                border-radius: 12px;
                padding: 16px;
                /* Fixed height — both boxes identical so they align */
                height: 340px;
                position: relative;
                overflow: hidden;        /* clip canvas if it tries to overflow */
            }}

            /* Canvas fills 100% of the box, nothing more */
            .chart-box canvas {{
                position: absolute;
                inset: 16px;             /* matches padding so canvas fits inside */
                width: calc(100% - 32px) !important;
                height: calc(100% - 32px) !important;
            }}

            /* ── Export buttons ── */
            .export-row {{
                display: flex; gap: 10px; flex-wrap: wrap;
                animation: fadeDown .4s .2s ease both;
            }}

            /* ── Search bar ── */
            .search-row {{
                display: flex; gap: 10px; flex-wrap: wrap; align-items: center;
                animation: fadeDown .4s .25s ease both;
            }}

            .input-wrap {{
                position: relative; flex: 1; min-width: 200px; max-width: 320px;
            }}

            .input-wrap input {{
                width: 100%;
                padding: 10px 14px 10px 36px;
                background: var(--surface); border: 1px solid var(--border);
                border-radius: 8px; color: var(--text);
                font-family: var(--body); font-size: 13px;
                outline: none; transition: border-color .2s;
            }}

            .input-wrap input:focus {{ border-color: var(--accent); }}

            .input-wrap .search-icon {{
                position: absolute; left: 12px; top: 50%;
                transform: translateY(-50%); color: var(--muted);
                font-size: 14px; pointer-events: none;
            }}

            select {{
                padding: 10px 14px;
                background: var(--surface); border: 1px solid var(--border);
                border-radius: 8px; color: var(--text);
                font-family: var(--body); font-size: 13px;
                outline: none; cursor: pointer; transition: border-color .2s;
            }}

            select:focus {{ border-color: var(--accent); }}

            .btn {{
                padding: 10px 20px; border-radius: 8px; border: none;
                font-family: var(--mono); font-size: 12px;
                font-weight: 700; letter-spacing: .05em;
                cursor: pointer; transition: all .2s; white-space: nowrap;
                display: inline-flex; align-items: center; gap: 6px;
            }}

            .btn-primary {{ background: var(--accent); color: #000; }}
            .btn-primary:hover {{
                background: #00f5ff;
                box-shadow: 0 0 20px rgba(0,217,232,.4);
            }}

            .btn-outline {{
                background: transparent;
                border: 1px solid var(--border); color: var(--muted);
            }}
            .btn-outline:hover {{ border-color: var(--accent); color: var(--accent); }}

            .btn-green {{
                background: transparent;
                border: 1px solid #2ea043; color: var(--accent2);
            }}
            .btn-green:hover {{
                background: rgba(126,231,135,.1);
                box-shadow: 0 0 16px rgba(126,231,135,.25);
            }}

            /* ── Table ── */
            .table-wrap {{
                background: var(--surface); border: 1px solid var(--border);
                border-radius: 12px; overflow: hidden;
                animation: fadeDown .4s .3s ease both;
            }}

            .table-header {{
                display: flex; align-items: center;
                justify-content: space-between;
                padding: 14px 20px; border-bottom: 1px solid var(--border);
            }}

            .table-title {{
                font-family: var(--mono); font-size: 12px;
                color: var(--muted); letter-spacing: .08em; text-transform: uppercase;
            }}

            .row-count {{
                font-family: var(--mono); font-size: 11px;
                background: var(--surface2); border: 1px solid var(--border);
                padding: 3px 10px; border-radius: 20px; color: var(--accent);
            }}

            table {{ width: 100%; border-collapse: collapse; }}

            th {{
                background: var(--surface2); padding: 11px 16px;
                text-align: left;
                font-family: var(--mono); font-size: 11px;
                letter-spacing: .08em; text-transform: uppercase;
                color: var(--muted); border-bottom: 1px solid var(--border);
                position: sticky; top: 0; z-index: 10;
            }}

            td {{
                padding: 12px 16px;
                border-bottom: 1px solid rgba(33,38,45,.6);
                font-size: 13px; vertical-align: middle;
            }}

            tr:last-child td {{ border-bottom: none; }}
            tr:hover td {{ background: var(--surface2); }}

            td a {{
                color: var(--text); text-decoration: none;
                transition: color .15s;
                display: -webkit-box;
                -webkit-line-clamp: 1; -webkit-box-orient: vertical;
                overflow: hidden; max-width: 600px;
            }}

            td a:hover {{ color: var(--accent); }}

            .source-pill {{
                display: inline-flex; align-items: center; gap: 5px;
                padding: 3px 10px; border-radius: 20px;
                font-family: var(--mono); font-size: 10px;
                letter-spacing: .06em; font-weight: 700; white-space: nowrap;
            }}

            .pill-hn  {{ background:rgba(255,100,0,.15); color:#ff6400; border:1px solid rgba(255,100,0,.3); }}
            .pill-rd  {{ background:rgba(255,69,0,.15);  color:#ff4500; border:1px solid rgba(255,69,0,.3); }}
            .pill-th  {{ background:rgba(0,217,232,.1);  color:var(--accent); border:1px solid rgba(0,217,232,.25); }}
            .pill-et  {{ background:rgba(126,231,135,.1); color:var(--accent2); border:1px solid rgba(126,231,135,.25); }}
            .pill-ys  {{ background:rgba(247,129,102,.1); color:var(--warn); border:1px solid rgba(247,129,102,.25); }}
            .pill-def {{ background:rgba(255,255,255,.06); color:var(--muted); border:1px solid var(--border); }}

            .ts {{ font-family: var(--mono); font-size: 11px; color: var(--muted); white-space: nowrap; }}

            .empty {{
                text-align: center; padding: 60px 20px;
                color: var(--muted); font-family: var(--mono); font-size: 12px;
            }}

            ::-webkit-scrollbar {{ width: 6px; height: 6px; }}
            ::-webkit-scrollbar-thumb {{ background: #30363d; border-radius: 6px; }}
            ::-webkit-scrollbar-track {{ background: transparent; }}

            @media (max-width: 900px) {{
                .shell {{ grid-template-columns: 1fr; }}
                .sidebar {{ display: none; }}
                .main {{ padding: 20px; gap: 20px; }}
                .stats-grid {{ grid-template-columns: 1fr 1fr; }}
                .charts-row {{ grid-template-columns: 1fr; }}
                .chart-box {{ height: 280px; }}
            }}

            @media (max-width: 600px) {{
                .stats-grid {{ grid-template-columns: 1fr; }}
                .input-wrap {{ max-width: 100%; }}
            }}

        </style>
    </head>
    <body>
    <div class="shell">

        <aside class="sidebar">
            <div class="sidebar-logo">
                <span class="logo-dot"></span>
                SCRAPER / v3
            </div>

            <span class="nav-label">Navigation</span>
            <a href="/" class="nav-link active"><span class="icon">⬛</span> Dashboard</a>
            <a href="/articles" class="nav-link"><span class="icon">📄</span> Articles API</a>
            <a href="/stats" class="nav-link"><span class="icon">📊</span> Stats API</a>
            <a href="/health" class="nav-link"><span class="icon">❤️</span> Health</a>

            <span class="nav-label">Export</span>
            <a href="/export/csv"   class="nav-link"><span class="icon">📥</span> Export CSV</a>
            <a href="/export/excel" class="nav-link"><span class="icon">📗</span> Export Excel</a>

            <div class="sidebar-footer">
                <div class="refresh-badge">
                    <span class="refresh-dot"></span>
                    AUTO-REFRESH 30s
                </div>
            </div>
        </aside>

        <main class="main">

            <div class="header">
                <div>
                    <div class="page-title">WEB <span>SCRAPER</span></div>
                    <div class="page-sub">Real-time article monitoring dashboard</div>
                </div>
                <div id="clock"></div>
            </div>

            <div class="stats-grid">
                <div class="stat-card">
                    <div class="stat-icon">📄</div>
                    <div class="stat-label">Total Articles</div>
                    <div class="stat-value accent">{len(df)}</div>
                </div>
                <div class="stat-card">
                    <div class="stat-icon">🌐</div>
                    <div class="stat-label">Active Sources</div>
                    <div class="stat-value">5</div>
                </div>
                <div class="stat-card">
                    <div class="stat-icon">✅</div>
                    <div class="stat-label">System Status</div>
                    <div class="stat-value green small">ACTIVE</div>
                </div>
            </div>

            <div class="analytics-card">
                <div class="analytics-title">📈 Source Analytics</div>
                <div class="charts-row">
                    <div class="chart-box">
                        <canvas id="sourceChart"></canvas>
                    </div>
                    <div class="chart-box">
                        <canvas id="pieChart"></canvas>
                    </div>
                </div>
            </div>

            <div class="export-row">
                <button class="btn btn-outline" onclick="window.location.href='/export/csv'">
                    📥 Export CSV
                </button>
                <button class="btn btn-green" onclick="window.location.href='/export/excel'">
                    📗 Export Excel
                </button>
            </div>

            <div class="search-row">
                <div class="input-wrap">
                    <span class="search-icon">🔍</span>
                    <input type="text" id="searchInput" placeholder="Search articles…">
                </div>
                <select id="sourceFilter">
                    <option value="">All Sources</option>
                    <option value="Hacker News">Hacker News</option>
                    <option value="Reddit">Reddit</option>
                    <option value="The Hindu">The Hindu</option>
                    <option value="Economic Times">Economic Times</option>
                    <option value="YourStory">YourStory</option>
                </select>
                <button class="btn btn-primary" onclick="searchArticles()">SEARCH</button>
                <button class="btn btn-outline"  onclick="resetTable()">RESET</button>
            </div>

            <div class="table-wrap">
                <div class="table-header">
                    <span class="table-title">Latest Articles</span>
                    <span class="row-count" id="rowCount">{len(df)} rows</span>
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>Title</th>
                            <th>Source</th>
                            <th>Scraped At</th>
                        </tr>
                    </thead>
                    <tbody id="tableBody">
                        {rows_html}
                    </tbody>
                </table>
            </div>

        </main>
    </div>

    <script>

        /* Clock */
        function updateClock() {{
            document.getElementById("clock").innerText =
                new Date().toLocaleString('en-US', {{ hour12: false }});
        }}
        setInterval(updateClock, 1000);
        updateClock();

        /* Source pill map */
        const PILL = {{
            "Hacker News":    "pill-hn",
            "Reddit":         "pill-rd",
            "The Hindu":      "pill-th",
            "Economic Times": "pill-et",
            "YourStory":      "pill-ys",
        }};
        function pillClass(src) {{ return PILL[src] || "pill-def"; }}

        /* Render rows */
        function renderRows(data) {{
            const tbody = document.getElementById("tableBody");
            document.getElementById("rowCount").innerText = data.length + " rows";
            if (!data.length) {{
                tbody.innerHTML = `<tr><td colspan="3"><div class="empty">No articles found.</div></td></tr>`;
                return;
            }}
            tbody.innerHTML = data.map(r => `
                <tr>
                    <td><a href="${{r.link}}" target="_blank" title="${{r.title}}">${{r.title}}</a></td>
                    <td><span class="source-pill ${{pillClass(r.source)}}">${{r.source}}</span></td>
                    <td><span class="ts">${{r.scraped_at}}</span></td>
                </tr>
            `).join("");
        }}

        async function searchArticles() {{
            const keyword = document.getElementById("searchInput").value.trim();
            const source  = document.getElementById("sourceFilter").value;
            let url = "/articles";
            if (keyword) url = `/search?keyword=${{encodeURIComponent(keyword)}}`;
            if (source)  url = `/source/${{encodeURIComponent(source)}}`;
            try {{ renderRows(await (await fetch(url)).json()); }}
            catch(e) {{ console.error(e); }}
        }}

        async function resetTable() {{
            document.getElementById("searchInput").value = "";
            document.getElementById("sourceFilter").value = "";
            try {{ renderRows(await (await fetch("/articles")).json()); }}
            catch(e) {{ console.error(e); }}
        }}

        document.getElementById("searchInput")
            .addEventListener("keydown", e => {{ if (e.key === "Enter") searchArticles(); }});

        /* Charts */
        async function loadCharts() {{
            try {{
                const data   = await (await fetch("/stats")).json();
                const labels = data.source_breakdown.map(i => i.source);
                const counts = data.source_breakdown.map(i => i.count);
                const colors = ['#00adb5','#ff5722','#ffc107','#4caf50','#9c27b0','#03a9f4'];

                /* Bar chart */
                new Chart(document.getElementById('sourceChart'), {{
                    type: 'bar',
                    data: {{
                        labels,
                        datasets: [{{
                            label: 'Articles',
                            data: counts,
                            backgroundColor: colors,
                            borderRadius: 6,
                            borderWidth: 0
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{ legend: {{ display: false }} }},
                        scales: {{
                            y: {{ beginAtZero: true, grid: {{ color: 'rgba(0,0,0,0.06)' }} }},
                            x: {{ grid: {{ display: false }} }}
                        }}
                    }}
                }});

                /* Doughnut chart
                   - maintainAspectRatio: false  → respects container height
                   - layout.padding            → breathing room inside the white box
                   - legend position: 'right'  → avoids legend eating vertical space
                     and cutting off the chart
                */
                new Chart(document.getElementById('pieChart'), {{
                    type: 'doughnut',
                    data: {{
                        labels,
                        datasets: [{{
                            data: counts,
                            backgroundColor: colors,
                            borderWidth: 3,
                            borderColor: '#ffffff',
                            hoverOffset: 8
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        cutout: '58%',
                        layout: {{
                            padding: {{ top: 8, bottom: 8, left: 8, right: 8 }}
                        }},
                        plugins: {{
                            legend: {{
                                position: 'right',
                                align: 'center',
                                labels: {{
                                    boxWidth: 12,
                                    boxHeight: 12,
                                    padding: 16,
                                    color: '#333',
                                    font: {{ size: 12 }}
                                }}
                            }}
                        }}
                    }}
                }});

            }} catch(e) {{ console.error(e); }}
        }}

        loadCharts();
        setInterval(() => {{ location.reload(); }}, 30000);

    </script>
    </body>
    </html>
    """

    return HTMLResponse(html)

@app.get("/articles")
async def get_articles():

    cached = get_cached_data(
        "latest_articles"
    )

    if cached:

        return cached

    df = fetch_df(
        """
        SELECT
            title,
            link,
            source,
            scraped_at
        FROM articles
        ORDER BY scraped_at DESC
        LIMIT 100
        """
    )

    result = df.to_dict(
        orient="records"
    )

    cache_data(
        "latest_articles",
        result
    )
    return result

@app.get("/search")
async def search_articles(keyword: str):
    df = fetch_df(
        "SELECT title,link,source,scraped_at FROM articles WHERE title LIKE ? ORDER BY scraped_at DESC LIMIT 50",
        params=(f"%{keyword}%",)
    )
    return df.to_dict(orient="records")

@app.get("/source/{source_name}")
async def get_by_source(source_name: str):
    df = fetch_df(
        "SELECT title,link,source,scraped_at FROM articles WHERE source=? ORDER BY scraped_at DESC LIMIT 100",
        params=(source_name,)
    )
    return df.to_dict(orient="records")

@app.get("/stats")
async def get_stats():
    total      = fetch_df("SELECT COUNT(*) as count FROM articles")
    sources    = fetch_df("SELECT source, COUNT(*) as count FROM articles GROUP BY source ORDER BY count DESC")
    latest     = fetch_df("SELECT scraped_at FROM articles ORDER BY scraped_at DESC LIMIT 1")
    top_source = fetch_df("SELECT source, COUNT(*) as count FROM articles GROUP BY source ORDER BY count DESC LIMIT 1")
    return {
        "status":           "success",
        "total_articles":   int(total['count'][0]),
        "total_sources":    len(sources),
        "latest_scrape":    latest['scraped_at'][0] if not latest.empty else "N/A",
        "top_source":       top_source['source'][0] if not top_source.empty else "N/A",
        "top_source_count": int(top_source['count'][0]) if not top_source.empty else 0,
        "source_breakdown": sources.to_dict(orient="records"),
    }

@app.get("/health")
async def health():
    return {"status": "UP", "service": "Smart Web Scraper System"}

@app.get("/export/csv")
async def export_csv():
    df = fetch_df("SELECT * FROM articles")
    return Response(
        content=df.to_csv(index=False),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=articles.csv"},
    )

@app.get("/export/excel")
async def export_excel():
    df = fetch_df("SELECT * FROM articles ORDER BY scraped_at DESC")

    wb = Workbook()

    # ── Sheet 1: Articles ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Articles"

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="0D7C85")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),  bottom=Side(style="thin", color="D0D0D0"),
    )

    headers = list(df.columns)
    for ci, col in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=col.upper())
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = thin

    alt_fill  = PatternFill("solid", fgColor="F0FAFA")
    norm_fill = PatternFill("solid", fgColor="FFFFFF")
    data_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0D7C85", underline="single")

    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = alt_fill if ri % 2 == 0 else norm_fill
        for ci, value in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
            col_key = headers[ci - 1].lower()
            if col_key == "link" and value:
                cell.hyperlink = str(value); cell.font = link_font
            elif col_key == "source":
                clr = SOURCE_COLORS_XLSX.get(str(value), "555555")
                cell.font = Font(name="Arial", size=10, bold=True, color=clr)
            else:
                cell.font = data_font
            cell.fill = fill

    col_widths = {"title": 60, "link": 50, "source": 18, "scraped_at": 22}
    for ci, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col.lower(), 20)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Sheet 2: Summary ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    sources_df = fetch_df(
        "SELECT source, COUNT(*) as article_count FROM articles GROUP BY source ORDER BY article_count DESC"
    )

    ws2["A1"] = "SOURCE BREAKDOWN"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="0D7C85")

    for ci, hdr in enumerate(["Source", "Article Count", "Share (%)"], 1):
        c = ws2.cell(row=3, column=ci, value=hdr)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = thin

    for i, srow in enumerate(sources_df.itertuples(index=False), 4):
        ws2.cell(row=i, column=1, value=srow.source).border = thin
        ws2.cell(row=i, column=1).font = Font(name="Arial", size=11)
        cc = ws2.cell(row=i, column=2, value=srow.article_count)
        cc.border = thin; cc.font = Font(name="Arial", size=11)
        cc.alignment = Alignment(horizontal="center")
        sc = ws2.cell(row=i, column=3, value=f"=B{i}/SUM(B4:B{3+len(sources_df)})")
        sc.border = thin; sc.number_format = "0.0%"
        sc.alignment = Alignment(horizontal="center")

    total_row = 4 + len(sources_df)
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=11)
    t = ws2.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})")
    t.font = Font(name="Arial", bold=True); t.border = thin
    ws2.cell(row=total_row, column=1).border = thin

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 14
    ws2.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=articles.xlsx"},
    )

@app.get("/export")
async def export_redirect():
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/export/csv")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
